import os
from openpyxl import Workbook, load_workbook

def initialize_workbook():
    if os.path.exists("Contacts.xlsx"):
        workbook = load_workbook("Contacts.xlsx")
        worksheet = workbook["Contacts"]
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Contacts"

        worksheet.cell(row=4, column=7).value = "Name"
        worksheet.cell(row=4, column=10).value = "Mobile Number"

        workbook.save("Contacts.xlsx")

        print("Contacts.xlsx not found.")
        print("New Contacts.xlsx file created successfully!\n")
    
    return workbook, worksheet

workbook, worksheet = initialize_workbook()

def search_logic(search_by = 0, search_value = 0):

    current_row = 6
    
    if(search_by == 1):
        current_column = 10
    else:
        current_column = 7
        
    while True:

        current_cell_value = worksheet.cell(row = current_row, column = current_column).value
        if(current_cell_value is None):
            return None
        if(current_cell_value == search_value):
            return current_row
        current_row += 2

def create_contact():
    print()
    print("="*30)
    print("NEW CONTACT".center(30))
    print("="*30)
    print()
    next_row = 6

    while True:
        if(worksheet.cell(row = next_row, column = 7).value is None):
            break
        next_row += 2
    try:
        contact_name = input("Enter Contact name   : ")
        contact_number = input("Enter Contact number : ")

        if(len(contact_number) != 10 or contact_number.isdigit() != True):
            print("Invalid mobile number!\n")
            return
    except:
        print("Invalid Input!\n")
        return
    
    name_exist = search_logic(search_by = 2, search_value = contact_name)
    number_exist = search_logic(search_by = 1, search_value = contact_number)

    if (name_exist is not None):
        print("Name already exists!\n")
    elif(number_exist is not None):
        print("Number already exists!\n")
    else:
        worksheet.cell(row = next_row, column = 7).value = contact_name
        worksheet.cell(row = next_row, column = 10).value = contact_number
        
        workbook.save("Contacts.xlsx")
        print("Contact added successfully.\n")

def search_contact():
    print()
    print("="*30)
    print("SEARCH CONTACTS".center(30))
    print("="*30)
    print()
    search_by = int(input("Want to search by?\n1. Number\n2. Name\nEnter your choice: "))
    if(search_by ==  1):
        search_value = input("Enter Number: ")
    elif(search_by == 2):
        search_value = input("Enter Name: ")
    else:
        print("Invalid input!")
        return
    
    is_exist = search_logic(search_by, search_value)
    print(is_exist)
    if(is_exist is not None):
        print(f"\nName          : {worksheet.cell(row = is_exist, column = 7).value}\nMobile Number : {worksheet.cell(row = is_exist, column = 10).value}\n")
    else:
        print("Contact not found!")

def delete_contact():
    print()
    print("="*30)
    print("DELETE CONTACTS".center(30))
    print("="*30)
    print()
    search_by = int(input("Want to delete contact by?\n1. Number\n2. Name\nEnter your choice: "))
    if(search_by ==  1):
        search_value = input("Enter Number: ")
    elif(search_by == 2):
        search_value = input("Enter Name: ")
    else:
        print("Invalid input!\n")
        return
    
    is_exist = search_logic(search_by, search_value)

    if(is_exist is not None):
        
        print(f"\nName          : {worksheet.cell(row = is_exist, column = 7).value}\nMobile Number : {worksheet.cell(row = is_exist, column = 10).value}\n")
        
        delete_confirmation = input("Confirm to delete this contact?(y/n): ").lower()
        
        if(delete_confirmation == "y"):
            worksheet.delete_rows(is_exist, 2)
            workbook.save("Contacts.xlsx")
            print("Contact deleted successfully.\n")
        else:
            print("Deletion terminated!\n")

    else:
        print("Contact not found!\n")

def display_contacts():
    print()
    print("="*30)
    print("CONTACTS LIST".center(30))
    print("="*30)
    print()
    n = "Name"
    N = "Mobile Number"
    print(f"{n:<20}{N}")
    print("-"*35)

    current_row = 6

    while worksheet.cell(row = current_row, column = 7).value is not None:
        name = worksheet.cell(row = current_row, column = 7).value 
        number = worksheet.cell(row = current_row, column = 10).value
        print(f"{name:<20}{number}")
        current_row += 2
    print()

print()
print("="*50)
print("CONTACT MANAGMENT SYSTEM".center(50))
print("="*50)
print("\n1. Create Contact\n2. Delete Contact\n3. Display Contacts\n4. Search Contact\n5. Exit\n")

while(True):
    print("="*50)
    try:
        operation = int(input("Enter your choice: "))
    except (ValueError):
        print("Choose valid option!")
        continue
    
    match(operation):
        case 1:
            create_contact()
        case 2:
            delete_contact()
        case 3:
            display_contacts()
        case 4:
            search_contact()
        case 5:
            print("Thankyou!\n")
            break
        case _:
            print("Not a valid choice!\n")
            break
    try:
        input("Press enter to continue.")
    except:
        pass